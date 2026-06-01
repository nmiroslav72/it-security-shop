"""
Uniview scraper sa boma.rs
Pokretanje: python uniview_scraper.py
"""
from playwright.sync_api import sync_playwright
import json, re, openpyxl
from openpyxl.styles import Font, PatternFill

BASE_URL   = "https://boma.rs"
EMAIL      = "nmiroslav72@yahoo.com"
PASSWORD   = "verona72"
MULTIPLIER = 118 * 1.2 * 1.48

CATEGORIES = [
    ("HDCVI kamere",    BASE_URL + "/uniview/allArticles/uniview/HDCVI_Kamere",    "hd-kamere",             True),
    ("IP kamere",       BASE_URL + "/uniview/allArticles/uniview/IP_Kamere",        "ip-kamere",             True),
    ("HDCVI rekorderi", BASE_URL + "/uniview/allArticles/uniview/HDCVI_Rekorderi",  "dvr-digitalni-snimaci", False),
    ("NVR rekorderi",   BASE_URL + "/uniview/allArticles/uniview/NVR_Rekorderi",    "nvr-mrezni-snimaci",    False),
]

def make_slug(name):
    return re.sub(r'-+', '-', re.sub(r'[^a-z0-9]+', '-', name.lower())).strip('-')

def guess_rez(name):
    n = name.lower()
    if "8mp" in n or "4k" in n: return "8MPX"
    if "5mp" in n: return "5MPX"
    if "4mp" in n: return "4MPX"
    if "3mp" in n: return "3MPX"
    if "2mp" in n or "1080" in n: return "2MPX"
    return ""

def login(page):
    print("Logujem se...")
    page.goto(BASE_URL + "/auth/login", wait_until="networkidle")
    page.wait_for_timeout(2000)
    page.fill("input[name='email']", EMAIL)
    page.fill("input[name='password']", PASSWORD)
    page.locator("button[type='submit']").first.click()
    page.wait_for_timeout(4000)
    print(f"  URL: {page.url}")

def get_available_links(page, category_url):
    print(f"\n  Otvaram: {category_url}")
    page.goto(category_url, wait_until="networkidle")
    page.wait_for_timeout(3000)
    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    page.wait_for_timeout(2000)

    available = []
    items = page.locator("span.article-wrap").all()
    print(f"  Ukupno: {len(items)}")

    for item in items:
        try:
            btn = item.locator(".btn-dodaj-grid").first
            if btn.count() == 0: continue
            if "DODAJ U KORPU" not in btn.inner_text().strip().upper(): continue
            href = item.locator(".article-image a").get_attribute("href") or ""
            href = href.strip()
            if href:
                full = href if href.startswith("http") else BASE_URL + href
                available.append(full)
        except: continue

    print(f"  Na stanju: {len(available)}")
    return available

def scrape_product(page, url):
    page.goto(url, wait_until="networkidle")
    page.wait_for_timeout(2500)
    page.evaluate("window.scrollTo(0, 400)")
    page.wait_for_timeout(1000)

    # NAZIV - prva linija
    name = ""
    try:
        full = page.locator(".mob-title-category").first.inner_text()
        lines = [l.strip() for l in full.splitlines() if l.strip()]
        name = lines[0] if lines else ""
        if name and not name.startswith("Uniview"):
            name = "Uniview " + name
    except: pass

    # CENA EUR
    price_eur = None
    try:
        txt = page.locator(".item_price_rsd").first.inner_text().strip()
        m = re.search(r'([\d]+[.,][\d]+)', txt.replace(" ", ""))
        if m:
            price_eur = float(m.group(1).replace(",", "."))
    except: pass

    # OPIS - drugi red iz mob-title-category
    desc = ""
    try:
        full = page.locator(".mob-title-category").first.inner_text()
        lines = [l.strip() for l in full.splitlines() if l.strip()]
        if len(lines) > 1:
            desc = " ".join(lines[1:]).rstrip(";").strip()
    except: pass

    # Dodaj SPAN specifikacije
    try:
        spans = page.locator("span").all()
        specs = []
        for span in spans:
            txt = span.inner_text().strip()
            if 10 < len(txt) < 200 and "Kategorije" not in txt and "©" not in txt:
                specs.append(txt)
        if specs:
            desc = desc + " | " + " | ".join(specs[:8])
    except: pass

    # Dodaj tabelu iz #home
    try:
        rows = page.locator("#home tr").all()
        table_specs = []
        for row in rows[:15]:
            cells = row.locator("td").all()
            if len(cells) >= 2:
                key = cells[0].inner_text().strip()
                val = cells[1].inner_text().strip()
                if key and val:
                    table_specs.append(key + ": " + val)
        if table_specs:
            desc = desc + " || " + " | ".join(table_specs)
    except: pass

    desc = desc.strip()[:700]

    # SLIKA - samo /img/articles/ sa _01.
    image = ""
    try:
        imgs = page.locator("img").all()
        for img in imgs:
            src = img.get_attribute("src") or ""
            if "/img/articles/" in src and "_01." in src:
                image = src if src.startswith("http") else BASE_URL + src
                break
    except: pass

    return {
        "name": name, "url": url,
        "price_eur": price_eur,
        "price_rsd": round(price_eur * MULTIPLIER) if price_eur else None,
        "desc": desc, "image": image,
    }

def main():
    print("=" * 60)
    print("UNIVIEW SCRAPER - boma.rs")
    print(f"Cena: x{MULTIPLIER:.3f}")
    print("=" * 60)

    all_products = []
    wooId = 22001
    seen_slugs = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=30)
        page = browser.new_context(viewport={"width": 1280, "height": 900}).new_page()

        try:
            login(page)
            page.wait_for_timeout(2000)

            for cat_name, cat_url, cat_slug, do_rez in CATEGORIES:
                print(f"\n{'='*50}\n{cat_name} -> {cat_slug}")
                links = get_available_links(page, cat_url)

                for i, link in enumerate(links, 1):
                    print(f"  [{i}/{len(links)}] ", end="", flush=True)
                    try:
                        prod = scrape_product(page, link)
                    except Exception as e:
                        print(f"GRESKA: {e}")
                        continue

                    if not prod or not prod["name"]:
                        print("nema naziva")
                        continue

                    rez = guess_rez(prod["name"]) if do_rez else ""
                    slug = make_slug(prod["name"])
                    if slug in seen_slugs:
                        slug = f"{slug}-{wooId}"
                    seen_slugs.append(slug)

                    all_products.append({
                        **prod, "wooId": wooId, "slug": slug,
                        "cat_slug": cat_slug, "cat_name": cat_name, "rezolucija": rez,
                    })
                    wooId += 1
                    print(f"OK {prod['name'][:40]} EUR={prod['price_eur']} img={'DA' if prod['image'] else 'NE'}")
                    page.wait_for_timeout(800)

        finally:
            browser.close()

    if not all_products:
        print("Nema proizvoda!")
        return

    print(f"\nUKUPNO: {len(all_products)}")

    with open("uniview_products.json", "w", encoding="utf-8") as f:
        json.dump(all_products, f, ensure_ascii=False, indent=2)
    print("uniview_products.json sacuvan")

    # Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Uniview"
    headers = ["wooId","Naziv","Kategorija","Rezolucija","EUR","RSD","Slika","Opis","URL"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1d3eb8")
    for r, prod in enumerate(all_products, 2):
        ws.cell(r,1,prod["wooId"]); ws.cell(r,2,prod["name"])
        ws.cell(r,3,prod["cat_name"]); ws.cell(r,4,prod["rezolucija"])
        ws.cell(r,5,prod["price_eur"]); ws.cell(r,6,prod["price_rsd"])
        ws.cell(r,7,prod["image"]); ws.cell(r,8,(prod["desc"] or "")[:200])
        ws.cell(r,9,prod["url"])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 2, 60)
    wb.save("uniview_products.xlsx")
    print("uniview_products.xlsx sacuvan")

    # Seed blok - lokalne putanje za slike
    lines = [
        "  // UNIVIEW - boma.rs",
        f"  // EUR x {MULTIPLIER:.3f}",
    ]
    for prod in all_products:
        if not prod["price_rsd"]: continue
        name = prod["name"].replace("'", "\\'")
        desc = (prod["desc"] or "").replace("'", "\\'").replace('"', '\\"')[:500]
        if prod["image"] and "/img/articles/" in prod["image"]:
            fname = prod["image"].split("/")[-1]
            imgs = f"['/uploads/uniview/{fname}']"
        elif prod["image"]:
            imgs = f"['{prod['image']}']"
        else:
            imgs = "[]"
        attrs = f'{{"brand":"Uniview","rezolucija":"{prod["rezolucija"]}"}}' if prod["rezolucija"] else '{"brand":"Uniview"}'
        lines.append(
            f"  {{ wooId:{prod['wooId']}, name:'{name}', slug:'{prod['slug']}',\n"
            f"    description:'{desc}',\n"
            f"    brand:'Uniview', price:{prod['price_rsd']}.0, regularPrice:{prod['price_rsd']}.0, salePrice:{prod['price_rsd']}.0,\n"
            f"    showPrice:true, inStock:true,\n"
            f"    images:{imgs}, attributes:{attrs}, categorySlug:'{prod['cat_slug']}' }},"
        )
    with open("uniview_seed_block.txt", "w", encoding="utf-8") as f:
        f.write('\n'.join(lines))
    print("uniview_seed_block.txt sacuvan")
    print("\nGotovo!")

if __name__ == "__main__":
    main()
