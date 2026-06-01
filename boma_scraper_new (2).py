"""
Boma.rs scraper - Imou oprema
- Sa liste uzima samo proizvode gde dugme kaze "DODAJ U KORPU"
- Za svaki takav proizvod ulazi i uzima sliku, opis, cenu
Pokretanje: python "boma scraper new.py"
"""
from playwright.sync_api import sync_playwright
import json, re, openpyxl
from openpyxl.styles import Font, PatternFill

BASE_URL    = "https://boma.rs"
EMAIL       = "nmiroslav72@yahoo.com"
PASSWORD    = "verona72"
MULTIPLIER  = 118 * 1.2 * 1.48  # = 209.664

CATEGORIES = [
    ("IP kamere",     BASE_URL + "/imou/allArticles/imou/IP_Kamere",     "ip-kamere",          True),
    ("NVR rekorderi", BASE_URL + "/imou/allArticles/imou/NVR_Rekorderi", "nvr-mrezni-snimaci", False),
]

def make_slug(name):
    return re.sub(r'-+', '-', re.sub(r'[^a-z0-9]+', '-', name.lower())).strip('-')

def guess_rez(name):
    n = name.lower()
    if "8mp" in n or "4k" in n: return "8MPX"
    if "5mp" in n: return "5MPX"
    if "4mp" in n: return "4MPX"
    if "3mp" in n: return "3MPX"
    if "2mp" in n or "1080" in n: return "2MPX"
    return ""

def login(page):
    print("🔐 Logujem se...")
    page.goto(BASE_URL + "/auth/login", wait_until="networkidle")
    page.wait_for_timeout(2000)
    page.fill("input[name='email']", EMAIL)
    page.fill("input[name='password']", PASSWORD)
    page.locator("button[type='submit']").first.click()
    page.wait_for_timeout(4000)
    print(f"  ✅ URL: {page.url}")

def get_available_links(page, category_url):
    """Sa liste uzima samo linkove gde dugme kaze 'DODAJ U KORPU'"""
    print(f"\n  Otvaram listu: {category_url}")
    page.goto(category_url, wait_until="networkidle")
    page.wait_for_timeout(3000)
    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    page.wait_for_timeout(2000)

    available = []
    items = page.locator("span.article-wrap").all()
    print(f"  Ukupno artikala: {len(items)}")

    for item in items:
        try:
            # Proveri tekst dugmeta
            btn = item.locator(".btn-dodaj-grid").first
            if btn.count() == 0:
                continue
            btn_text = btn.inner_text().strip().upper()
            if "DODAJ U KORPU" not in btn_text:
                continue  # Preskoci "TRENUTNO NEDOSTUPNO"

            # Uzmi link
            href = item.locator(".article-image a").get_attribute("href") or ""
            href = href.strip()
            if href:
                full = href if href.startswith("http") else BASE_URL + href
                available.append(full)
        except:
            continue

    print(f"  Na stanju (ima 'DODAJ U KORPU'): {len(available)}")
    return available

def scrape_product(page, url):
    """Ulazi u stranicu proizvoda i uzima sve podatke"""
    page.goto(url, wait_until="networkidle")
    page.wait_for_timeout(2500)
    page.evaluate("window.scrollTo(0, 400)")
    page.wait_for_timeout(500)

    # NAZIV - .mob-title-category
    name = ""
    try:
        name = page.locator(".mob-title-category").first.inner_text()
        name = re.sub(r'\s+', ' ', name).strip().rstrip(';').strip()
    except: pass

    # CENA EUR - .item_price_rsd sadrzi EUR vrednost (npr "40,85  €")
    price_eur = None
    try:
        txt = page.locator(".item_price_rsd").first.inner_text().strip()
        m = re.search(r'([\d]+[.,][\d]+)', txt.replace(" ", ""))
        if m:
            price_eur = float(m.group(1).replace(",", "."))
    except: pass

    # OPIS - .article-info ili kratki opis sa stranice
    desc = ""
    try:
        desc = page.locator(".article-info p").first.inner_text().strip()
    except: pass
    if not desc:
        try:
            desc = page.locator(".mob-title-category").first.inner_text().strip()
        except: pass

    # SLIKA - tražimo prvu img koja nije logo/ikona
    image = ""
    # Najpre probaj direktno iz poznate putanje (boma koristi /public/img/imou/MODEL_01.jpg)
    try:
        imgs = page.locator("img").all()
        for img in imgs:
            src = img.get_attribute("src") or ""
            if not src: continue
            if any(x in src.lower() for x in ["logo", "cart", "dahua", "icon", "flag", "svg", "banner"]):
                continue
            if re.search(r'\.(jpg|jpeg|png|webp)', src.lower()):
                image = src if src.startswith("http") else BASE_URL + src
                break
    except: pass

    # Ako nije nađena slika, traži background-image
    if not image:
        try:
            result = page.evaluate("""
                () => {
                    const els = document.querySelectorAll('[style*="background-image"]');
                    for (const el of els) {
                        const m = (el.getAttribute('style')||'').match(/url\\(['"]?([^'"\\)]+)['"]?\\)/);
                        if (m && m[1] && !m[1].includes('logo') && !m[1].includes('icon')) return m[1];
                    }
                    return '';
                }
            """)
            if result:
                image = result if result.startswith("http") else BASE_URL + result
        except: pass

    return {
        "name": name,
        "url": url,
        "price_eur": price_eur,
        "price_rsd": round(price_eur * MULTIPLIER) if price_eur else None,
        "desc": desc[:600],
        "image": image,
    }

def main():
    print("=" * 60)
    print("BOMA.RS SCRAPER - Imou (samo na stanju)")
    print(f"Cena: EUR × 118 × 1.2 × 1.48 = ×{MULTIPLIER:.3f}")
    print("=" * 60)

    all_products = []
    wooId = 21001
    seen_slugs = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=30)
        context = browser.new_context(viewport={"width": 1280, "height": 900})
        page = context.new_page()

        try:
            login(page)
            page.wait_for_timeout(2000)

            for cat_name, cat_url, cat_slug, do_rez in CATEGORIES:
                print(f"\n{'='*50}")
                print(f"📦 {cat_name}")

                # Korak 1: sa liste uzmi linkove na stanju
                links = get_available_links(page, cat_url)

                # Korak 2: uđi u svaki i uzmi podatke
                for i, link in enumerate(links, 1):
                    print(f"  [{i}/{len(links)}] ", end="", flush=True)
                    try:
                        prod = scrape_product(page, link)
                    except Exception as e:
                        print(f"❌ {e}")
                        continue

                    if not prod or not prod["name"]:
                        print(f"⚠️  nema naziva")
                        continue

                    rez  = guess_rez(prod["name"]) if do_rez else ""
                    slug = make_slug(prod["name"])
                    if slug in seen_slugs:
                        slug = f"{slug}-{wooId}"
                    seen_slugs.append(slug)

                    all_products.append({
                        **prod,
                        "wooId": wooId, "slug": slug,
                        "cat_slug": cat_slug, "cat_name": cat_name,
                        "rezolucija": rez,
                    })
                    wooId += 1
                    print(f"✅ {prod['name'][:45]} | EUR={prod['price_eur']} | RSD={prod['price_rsd']}")
                    page.wait_for_timeout(800)

        finally:
            browser.close()

    if not all_products:
        print("\n❌ Nije pronađen nijedan proizvod!")
        return

    print(f"\n\n✅ UKUPNO: {len(all_products)} proizvoda")

    # JSON
    with open("imou_products.json", "w", encoding="utf-8") as f:
        json.dump(all_products, f, ensure_ascii=False, indent=2)
    print("📄 imou_products.json")

    # Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Imou"
    headers = ["wooId","Naziv","Kategorija","Rezolucija","EUR","RSD","Slika","Opis","URL"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1d3eb8")
    for r, prod in enumerate(all_products, 2):
        ws.cell(r,1,prod["wooId"]); ws.cell(r,2,prod["name"])
        ws.cell(r,3,prod["cat_name"]); ws.cell(r,4,prod["rezolucija"])
        ws.cell(r,5,prod["price_eur"]); ws.cell(r,6,prod["price_rsd"])
        ws.cell(r,7,prod["image"]); ws.cell(r,8,(prod["desc"] or "")[:200])
        ws.cell(r,9,prod["url"])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 2, 60)
    wb.save("imou_products.xlsx")
    print("📊 imou_products.xlsx")

    # Seed blok
    lines = [
        "  // IMOU - boma.rs (samo na stanju)",
        f"  // EUR × 118 × 1.2 × 1.48 = ×{MULTIPLIER:.3f}",
    ]
    for prod in all_products:
        if not prod["price_rsd"]: continue
        name = prod["name"].replace("'", "\\'")
        desc = (prod["desc"] or "").replace("'", "\\'").replace('"', '\\"')[:500]
        imgs = f"['{prod['image']}']" if prod["image"] else "[]"
        attrs = f'{{"brand":"Imou","rezolucija":"{prod["rezolucija"]}"}}' if prod["rezolucija"] else '{"brand":"Imou"}'
        lines.append(
            f"  {{ wooId:{prod['wooId']}, name:'{name}', slug:'{prod['slug']}',\n"
            f"    description:'{desc}',\n"
            f"    brand:'Imou', price:{prod['price_rsd']}.0, regularPrice:{prod['price_rsd']}.0, salePrice:{prod['price_rsd']}.0,\n"
            f"    showPrice:true, inStock:true,\n"
            f"    images:{imgs}, attributes:{attrs}, categorySlug:'{prod['cat_slug']}' }},"
        )
    with open("imou_seed_block.txt", "w", encoding="utf-8") as f:
        f.write('\n'.join(lines))
    print("🌱 imou_seed_block.txt")
    print("\n🎉 Gotovo!")
    print("1. Proverite imou_products.xlsx")
    print("2. Dodajte imou_seed_block.txt u prisma/seed.ts")
    print("3. npx prisma db seed")

if __name__ == "__main__":
    main()
